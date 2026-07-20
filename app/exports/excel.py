from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.utils.grades import student_activity_mark, section_summary, course_total


def _style_header(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, num_cols, min_width=10, max_width=40):
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        longest = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(longest, max_width)


def _write_course_sheet(ws, course, student):
    ws.append([f"{course.code} — {course.title}", "", "", ""])
    ws.append([course.term, "", "", ""])
    ws.append([f"Student: {student.full_name} ({student.email})", "", "", ""])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(["Section", "Activity", "Obtained", "Total", "Percentage"])
    _style_header(ws, header_row, 5)

    for section in course.sections:
        for activity in section.activities:
            mark = student_activity_mark(student.id, activity)
            if mark is not None:
                obtained = mark.obtained_marks
                pct = f"{(obtained / activity.total_marks) * 100:.1f}%" if activity.total_marks else "—"
            else:
                obtained = "Not graded"
                pct = "—"
            ws.append([section.name, activity.name, obtained, activity.total_marks, pct])

    ws.append([])

    for section in course.sections:
        summary = section_summary(student.id, section)
        if summary:
            ws.append([
                f"{section.name} total",
                "",
                summary["obtained"],
                summary["possible"],
                f"{summary['percentage']:.1f}%",
            ])
        else:
            ws.append([f"{section.name} total", "", "Not graded", "", "—"])

    overall = course_total(student.id, course)
    ws.append([])
    if overall:
        ws.append(["Course total", "", "", "", f"{overall['current_percentage']:.1f}%"])
    else:
        ws.append(["Course total", "", "", "", "No sections yet"])

    _autofit_columns(ws, 5)


def build_course_workbook(course, student):
    """Single-course export for one student."""
    wb = Workbook()
    ws = wb.active
    ws.title = course.code[:31]  # Excel sheet name limit
    _write_course_sheet(ws, course, student)

    return _finalize(wb)


def build_multi_course_workbook(courses, student):
    """One sheet per course, for a student exporting all their results at once."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    if not courses:
        ws = wb.create_sheet("No courses")
        ws.append(["You are not enrolled in any published courses yet."])
    else:
        used_names = set()
        for course in courses:
            base_name = course.code[:31]
            name = base_name
            suffix = 1
            while name in used_names:
                suffix += 1
                name = f"{base_name[:28]}_{suffix}"
            used_names.add(name)

            ws = wb.create_sheet(name)
            _write_course_sheet(ws, course, student)

    return _finalize(wb)

def build_roster_workbook(course, students):
    """Staff-facing gradebook export: one row per student, one column per
    activity (grouped by section), plus section totals and course total."""
    wb = Workbook()
    ws = wb.active
    ws.title = course.code[:31]

    ws.append([f"{course.code} — {course.title} ({course.term})"])
    ws.append([])

    # Build the flat list of (section, activity) columns in order.
    activity_columns = []
    for section in course.sections:
        for activity in section.activities:
            activity_columns.append((section, activity))

    header_row_1 = ["Student", "Email"]
    header_row_2 = ["", ""]
    for section, activity in activity_columns:
        header_row_1.append(section.name)
        header_row_2.append(f"{activity.name} (/{activity.total_marks})")
    for section in course.sections:
        header_row_1.append("Section total")
        header_row_2.append(f"{section.name} (%)")
    header_row_1.append("")
    header_row_2.append("Course total (%)")

    ws.append(header_row_1)
    ws.append(header_row_2)
    header_start = ws.max_row - 1
    _style_header(ws, header_start, len(header_row_1))
    _style_header(ws, header_start + 1, len(header_row_2))

    for student in students:
        row = [student.full_name, student.email]

        for section, activity in activity_columns:
            mark = student_activity_mark(student.id, activity)
            row.append(mark.obtained_marks if mark is not None else "—")

        for section in course.sections:
            summary = section_summary(student.id, section)
            row.append(f"{summary['percentage']:.1f}" if summary else "—")

        overall = course_total(student.id, course)
        row.append(f"{overall['current_percentage']:.1f}" if overall else "—")

        ws.append(row)

    _autofit_columns(ws, len(header_row_1))
    return _finalize(wb)


def _finalize(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf